"""
SolidWorks BOM提取器（简化版 - 去除手动等待）
"""
import os
import time
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

try:
    import win32com.client
    import pythoncom
    import pywintypes
    PYWIN32_AVAILABLE = True
except ImportError:
    PYWIN32_AVAILABLE = False


@dataclass
class SWBOMItem:
    """SolidWorks BOM项（简化版）"""
    item_number: str
    part_name:   str
    quantity: int
    configuration: str = "Default"

    def to_dict(self) -> Dict[str, Any]:
        return {
            'item_number': self.item_number,
            'part_name': self.part_name,
            'quantity': self.quantity,
            'configuration': self.configuration
        }


class SolidWorksBOMExtractorOfficial:
    """SolidWorks BOM提取器（官方API - 简化版）"""

    def __init__(self):
        if not PYWIN32_AVAILABLE:
            raise ImportError("需要安装 pywin32: pip install pywin32")

        self.sw_app = None
        self.bom_items: List[SWBOMItem] = []
        self. assy_doc = None

    def connect_to_solidworks(self, max_retries: int = 10) -> bool:
        """
        连接到SolidWorks（自动化版本）

        Args:
            max_retries: 最大重试次数（默认10次，约10秒）
        """
        try:
            # ✅ 去除手动等待，直接初始化COM
            try:
                pythoncom.CoInitialize()
            except:
                pass

            print("🔌 正在连接到SolidWorks...")

            prog_ids = ["SldWorks.Application", "SolidWorks.Application"]

            # ✅ 快速重试，不显示过多信息
            for attempt in range(max_retries):
                for prog_id in prog_ids:
                    try:
                        self.sw_app = win32com.client.GetActiveObject(prog_id)

                        if self.sw_app:
                            print(f"✅ 成功连接到SolidWorks")
                            if self._verify_connection():
                                return True

                    except pywintypes.com_error as e:
                        if e.args[0] == -2147221021:
                            continue
                    except:
                        continue

                # 仅在重试时短暂等待
                if attempt < max_retries - 1:
                    time.sleep(1)

            # ✅ 连接失败，给出简洁提示
            print("❌ 无法连接到SolidWorks")
            print("   请确保：")
            print("   1. SolidWorks已打开")
            print("   2. 装配文件已在SolidWorks中打开")
            print("   3. 装配文件是当前活动文档")

            return False

        except Exception as e:
            print(f"❌ 连接失败: {e}")
            return False

    def _verify_connection(self) -> bool:
        """验证连接（自动化版本）"""
        try:
            active_doc = self.sw_app.ActiveDoc

            if not active_doc:
                print("⚠️ 未检测到活动文档")
                print("   请在SolidWorks中打开装配文件")
                return False

            print("✅ 检测到活动文档")

            doc_type = active_doc.GetType

            if doc_type != 2:  # 2 = 装配
                print(f"❌ 当前文档不是装配文件（类型={doc_type}）")
                print("   请在SolidWorks中打开. SLDASM文件")
                return False

            self.assy_doc = active_doc
            print("✅ 装配文档已就绪")

            # 快速检查组件数量
            try:
                components = self. assy_doc.GetComponents(False)
                if components:
                    print(f"   检测到 {len(components)} 个组件")
            except:
                pass

            return True

        except Exception as e:
            print(f"⚠️ 验证失败: {e}")
            return False

    def extract_bom(self, top_level_only: bool = False) -> List[SWBOMItem]:
        """提取BOM（简化版）"""
        if not self. assy_doc:
            raise RuntimeError("未连接到装配文档")

        print(f"\n{'='*70}")
        print(f"📖 提取BOM")
        print(f"{'='*70}\n")

        try:
            components = self.assy_doc. GetComponents(top_level_only)

            if not components:
                print("⚠️ 未找到组件")
                return []

            print(f"✅ 找到 {len(components)} 个组件")

            # 统计零件（按零件名去重）
            part_dict = {}

            for idx, comp in enumerate(components, 1):
                try:
                    comp_info = self._extract_component_info(comp, idx)

                    if comp_info:
                        # 使用零件名+配置作为唯一键
                        key = (comp_info['part_name'], comp_info['configuration'])

                        if key in part_dict:
                            part_dict[key]['quantity'] += 1
                        else:
                            part_dict[key] = comp_info

                except Exception as e:
                    continue

            # 转换为BOM项
            print(f"\n📊 生成BOM...")

            for idx, (key, info) in enumerate(sorted(part_dict.items()), start=1):
                bom_item = SWBOMItem(
                    item_number=str(idx),
                    part_name=info['part_name'],
                    quantity=info['quantity'],
                    configuration=info['configuration']
                )
                self.bom_items.append(bom_item)

            print(f"✅ BOM生成完成")
            print(f"   独特零件: {len(self.bom_items)}")
            print(f"   总数量: {sum(item.quantity for item in self.bom_items)}\n")

            return self.bom_items

        except Exception as e:
            print(f"❌ BOM提取失败: {e}")
            import traceback
            traceback.print_exc()
            raise

    def _extract_component_info(self, comp, idx: int) -> Optional[Dict[str, Any]]:
        """提取组件信息"""
        try:
            # 获取组件完整名称
            comp_full_name = f"Component_{idx}"

            try:
                if hasattr(comp, 'Name2'):
                    comp_full_name = str(comp.Name2)
            except:
                pass

            if comp_full_name == f"Component_{idx}":
                try:
                    if hasattr(comp, 'Name'):
                        comp_full_name = str(comp.Name)
                except:
                    pass

            # 提取零件名（去除实例号）
            part_name = comp_full_name
            if '-' in comp_full_name:
                parts = comp_full_name.rsplit('-', 1)
                if len(parts) == 2 and parts[1]. isdigit():
                    part_name = parts[0]

            # 配置
            config_name = "Default"
            try:
                if hasattr(comp, 'ReferencedConfiguration'):
                    config_name = str(comp.ReferencedConfiguration)
            except:
                pass

            return {
                'part_name': part_name,
                'quantity': 1,
                'configuration': config_name
            }

        except Exception as e:
            return None

    def export_to_excel(self, output_path: str):
        """导出到Excel（简化版）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("需要安装:   pip install openpyxl")

        print(f"📊 导出到Excel:   {os.path.basename(output_path)}")

        wb = Workbook()
        ws_bom = wb.active
        ws_bom.title = "BOM"

        # 表头
        headers = ['序号', '零件名称', '数量', '配置']
        ws_bom. append(headers)

        # 表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')

        for cell in ws_bom[1]:
            cell.fill = header_fill
            cell. font = header_font
            cell.alignment = header_alignment

        # 数据行
        for item in self.bom_items:
            ws_bom.append([
                item.item_number,
                item.part_name,
                item.quantity,
                item.configuration
            ])

        # 设置列宽
        ws_bom.column_dimensions['A'].width = 8
        ws_bom.column_dimensions['B'].width = 40
        ws_bom.column_dimensions['C'].width = 10
        ws_bom.column_dimensions['D'].width = 15

        # 数据行居中对齐
        for row in ws_bom. iter_rows(min_row=2, max_row=len(self.bom_items)+1):
            row[0].alignment = Alignment(horizontal='center')  # 序号
            row[2].alignment = Alignment(horizontal='center')  # 数量
            row[3].alignment = Alignment(horizontal='center')  # 配置

        # 统计工作表
        ws_stats = wb.create_sheet("统计")
        ws_stats.append(['统计项', '数值'])
        ws_stats.append(['独特零件种类', len(self.bom_items)])
        ws_stats.append(['总零件数量', sum(item.quantity for item in self.bom_items)])

        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        ws_stats.column_dimensions['A'].width = 20
        ws_stats.column_dimensions['B'].width = 15

        for row in ws_stats.iter_rows(min_row=2, max_row=3):
            row[1].alignment = Alignment(horizontal='center')

        wb.save(output_path)
        print(f"✅ 已导出到:  {os.path.basename(output_path)}\n")

    def export_to_json(self, output_path: str):
        """导出为JSON"""
        import json
        from datetime import datetime

        print(f"📊 导出到JSON:  {os.path.basename(output_path)}")

        bom_data = {
            'generated_at': datetime.now().isoformat(),
            'total_items': len(self.bom_items),
            'total_quantity': sum(item.quantity for item in self.bom_items),
            'items': [item. to_dict() for item in self.bom_items]
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(bom_data, f, ensure_ascii=False, indent=2)

        print(f"✅ 已导出到: {os.path.basename(output_path)}\n")

    def export_to_csv(self, output_path: str):
        """导出到CSV"""
        import csv

        print(f"📊 导出到CSV: {os. path.basename(output_path)}")

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 表头
            writer.writerow(['序号', '零件名称', '数量', '配置'])

            # 数据
            for item in self.bom_items:
                writer.writerow([
                    item.item_number,
                    item. part_name,
                    item.quantity,
                    item. configuration
                ])

            # 统计
            writer.writerow([])
            writer.writerow(['统计信息', '', '', ''])
            writer.writerow(['独特零件种类', len(self.bom_items), '', ''])
            writer.writerow(['总零件数量', sum(item.quantity for item in self.bom_items), '', ''])

        print(f"✅ 已导出到: {os.path. basename(output_path)}\n")

    def print_summary(self):
        """打印摘要"""
        print("="*70)
        print("📋 BOM摘要")
        print("="*70)

        total_quantity = sum(item.quantity for item in self.bom_items)

        print(f"独特零件:  {len(self.bom_items)}")
        print(f"总数量: {total_quantity}")

        print("\n" + "-" * 70)
        print(f"{'序号':<6} {'零件名称':<35} {'数量':<8} {'配置':<15}")
        print("-" * 70)

        for item in self.bom_items[: 20]:
            print(f"{item.item_number:<6} {item. part_name:<35} {item.quantity:<8} {item. configuration:<15}")

        if len(self.bom_items) > 20:
            print(f"... 还有 {len(self.bom_items) - 20} 项")

        print("="*70 + "\n")

    def disconnect(self):
        """断开连接"""
        try:
            pythoncom.CoUninitialize()
        except:
            pass

        self.sw_app = None
        self. assy_doc = None