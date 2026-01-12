"""
SolidWorks BOM自动提取器（无需打开SolidWorks）
尝试直接启动SolidWorks后台实例并打开文件
"""
import os
import time
from typing import List, Dict, Any, Optional
from dataclasses import dataclass

try:
    import win32com.client
    import pythoncom
    import pywintypes

    PYWIN32_AVAILABLE = True
except ImportError:
    PYWIN32_AVAILABLE = False


@dataclass
class SWBOMItem:
    """SolidWorks BOM项"""
    item_number: str
    part_name: str
    quantity: int
    configuration: str = "Default"
    material: str = ""
    mass: float = 0.0

    def to_dict(self) -> Dict[str, Any]:
        return {
            'item_number': self.item_number,
            'part_name': self.part_name,
            'quantity': self.quantity,
            'configuration': self.configuration,
            'material': self.material,
            'mass': self.mass
        }


class SolidWorksBOMExtractorAuto:
    """SolidWorks BOM自动提取器（后台模式）"""

    def __init__(self):
        if not PYWIN32_AVAILABLE:
            raise ImportError("需要安装 pywin32: pip install pywin32")

        self.sw_app = None
        self.bom_items: List[SWBOMItem] = []
        self.model_doc = None
        self.assy_doc = None  # ✅ 装配文档（与 model_doc 指向同一对象）
        self.started_solidworks = False

    def start_solidworks(self, visible: bool = False) -> bool:
        """
        启动SolidWorks实例（后台或可见）

        Args:
            visible: 是否显示SolidWorks窗口（False=后台运行）
        """
        print(f"\n🚀 启动SolidWorks {'(后台模式)' if not visible else '(可见模式)'}...")

        try:
            pythoncom.CoInitialize()

            # 方法1: 尝试连接现有实例
            try:
                print("   尝试连接到现有实例...")
                self.sw_app = win32com.client.GetActiveObject("SldWorks. Application")
                print("   ✓ 已连接到现有实例")
                self.started_solidworks = False

            except pywintypes.com_error:
                # 方法2: 启动新实例
                print("   启动新的SolidWorks实例...")

                try:
                    # ✅ 使用Dispatch启动新实例
                    self.sw_app = win32com.client.Dispatch("SldWorks.Application")
                    self.started_solidworks = True
                    print("   ✓ 已启动新实例")

                except Exception as e:
                    print(f"   ✗ Dispatch失败: {e}")

                    # 方法3: 使用DispatchEx（强制新实例）
                    try:
                        print("   尝试DispatchEx...")
                        self.sw_app = win32com.client.DispatchEx("SldWorks.Application")
                        self.started_solidworks = True
                        print("   ✓ DispatchEx成功")
                    except Exception as e2:
                        print(f"   ✗ DispatchEx失败: {e2}")
                        raise

            if not self.sw_app:
                raise RuntimeError("无法创建SolidWorks对象")

            # 设置可见性
            try:
                self.sw_app.Visible = visible
                print(f"   ✓ 可见性设置为: {visible}")
            except Exception as e:
                print(f"   ⚠ 设置可见性失败: {e}")

            # 等待SolidWorks完全启动
            if self.started_solidworks:
                print("   等待SolidWorks启动...")
                max_wait = 60  # 最多等60秒

                for i in range(max_wait):
                    try:
                        # 尝试访问一个简单的属性来验证是否就绪
                        _ = self.sw_app.Visible

                        # 等待至少5秒确保完全启动
                        if i >= 5:
                            print(f"   ✓ SolidWorks已就绪 ({i}秒)")
                            break
                    except:
                        pass

                    if i % 10 == 0 and i > 0:
                        print(f"   等待中... {i}/{max_wait}秒")

                    time.sleep(1)
                else:
                    print("   ⚠ 超时，但尝试继续...")

            print("✅ SolidWorks已准备就绪\n")
            return True

        except Exception as e:
            print(f"❌ 启动SolidWorks失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def open_assembly(self, sldasm_path: str) -> bool:
        """
        打开装配文件

        Args:
            sldasm_path: .  SLDASM文件的完整路径
        """
        if not self.sw_app:
            raise RuntimeError("SolidWorks未启动")

        if not os.path.exists(sldasm_path):
            raise FileNotFoundError(f"文件不存在: {sldasm_path}")

        abs_path = os.path.abspath(sldasm_path)
        print(f"📂 打开装配文件...")
        print(f"   {abs_path}")

        try:
            # swDocASSEMBLY = 2
            doc_type = 2

            # OpenDoc6 参数
            errors = pythoncom.Empty
            warnings = pythoncom.Empty

            print("   正在打开文件...")

            try:
                # 方法1: OpenDoc6（推荐）
                self.model_doc = self.sw_app.OpenDoc6(
                    abs_path,
                    doc_type,
                    1,  # swOpenDocOptions_Silent (静默打开)
                    "",
                    errors,
                    warnings
                )
            except Exception as e:
                print(f"   OpenDoc6失败: {e}")

                # 方法2: OpenDoc（简化版本）
                try:
                    print("   尝试OpenDoc...")
                    self.model_doc = self.sw_app.OpenDoc(abs_path, doc_type)
                except Exception as e2:
                    print(f"   OpenDoc失败: {e2}")
                    raise

            if not self.model_doc:
                raise RuntimeError("OpenDoc返回None，文件可能无法打开")

            # ✅ 关键：同时设置 assy_doc（extract_bom 使用这个）
            self.assy_doc = self.model_doc

            print("✅ 文件已打开\n")

            # 验证文档类型
            try:
                actual_type = self.model_doc.GetType
                if actual_type != 2:
                    print(f"⚠ 警告:  文档类型不是装配 (type={actual_type})")
            except:
                pass

            return True

        except Exception as e:
            print(f"❌ 打开文件失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def extract_bom(self, top_level_only: bool = False) -> List[SWBOMItem]:
        """提取BOM（完整修复版）"""
        if not self.assy_doc:
            raise RuntimeError("未连接到装配文档")

        print(f"{'=' * 70}")
        print(f"📖 提取BOM")
        print(f"{'=' * 70}\n")

        try:
            # ✅ 清空之前的BOM（防止累积）
            self.bom_items = []

            components = self.assy_doc.GetComponents(top_level_only)

            if not components:
                print("⚠️ 未找到组件")
                return []

            print(f"获取组件列表...")
            print(f"✓ 找到 {len(components)} 个组件\n")

            # 统计零件（按零件名+配置去重）
            part_dict = {}

            print("处理组件:")
            for idx, comp in enumerate(components, 1):
                try:
                    comp_info = self._extract_component_info(comp, idx)

                    if comp_info:
                        # 使用零件名+配置作为唯一键
                        key = (comp_info['part_name'], comp_info['configuration'])

                        if key in part_dict:
                            # 零件已存在，数量+1
                            part_dict[key]['quantity'] += 1
                        else:
                            # 新零件
                            part_dict[key] = comp_info

                    if idx % 5 == 0:
                        print(f"   已处理: {idx}/{len(components)}")

                except Exception as e:
                    continue

            # 最后一次进度显示
            print(f"   已处理: {len(components)}/{len(components)}")

            # 转换为BOM项
            print(f"\n生成BOM...")

            for idx, (key, info) in enumerate(sorted(part_dict.items()), start=1):
                bom_item = SWBOMItem(
                    item_number=str(idx),
                    part_name=info['part_name'],
                    quantity=info['quantity'],
                    configuration=info['configuration'],
                    material=info.get('material', ''),
                    mass=info.get('mass', 0.0)
                )
                self.bom_items.append(bom_item)

            print(f"\n✅ BOM提取完成")
            print(f"   独特零件: {len(self.bom_items)}")
            print(f"   总数量: {sum(item.quantity for item in self.bom_items)}\n")

            return self.bom_items

        except Exception as e:
            print(f"❌ BOM提取失败: {e}")
            import traceback
            traceback.print_exc()
            raise

    def _extract_component_info(self, comp, idx: int) -> Optional[Dict[str, Any]]:
        """提取组件信息"""
        try:
            # 组件名称
            comp_full_name = f"Component_{idx}"

            try:
                if hasattr(comp, 'Name2'):
                    comp_full_name = str(comp.Name2)
            except:
                pass

            if comp_full_name == f"Component_{idx}":
                try:
                    if hasattr(comp, 'Name'):
                        comp_full_name = str(comp.Name)
                except:
                    pass

            # 提取零件名（去除实例号）
            part_name = comp_full_name
            if '-' in comp_full_name:
                parts = comp_full_name.rsplit('-', 1)
                if len(parts) == 2 and parts[1].isdigit():
                    part_name = parts[0]

            # 配置
            config_name = "Default"
            try:
                if hasattr(comp, 'ReferencedConfiguration'):
                    config_name = str(comp.ReferencedConfiguration)
            except:
                pass

            # ✅ 尝试获取模型文档（用于提取材料和质量）
            material = ""
            mass = 0.0

            try:
                model = comp.GetModelDoc2()
                if model:
                    # 尝试获取材料
                    try:
                        mat = model.MaterialIdName
                        if mat:
                            material = str(mat)
                    except:
                        pass

                    # 尝试获取质量
                    try:
                        mass_prop = model.Extension.CreateMassProperty()
                        if mass_prop:
                            mass = float(mass_prop.Mass)
                    except:
                        pass
            except:
                # GetModelDoc2失败时忽略
                pass

            return {
                'part_name': part_name,
                'quantity': 1,
                'configuration': config_name,
                'material': material,
                'mass': mass
            }

        except Exception as e:
            return None

    def export_to_excel(self, output_path: str):
        """导出到Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            raise ImportError("需要安装:  pip install openpyxl")

        print(f"📊 导出到Excel:  {os.path.basename(output_path)}")

        wb = Workbook()
        ws = wb.active
        ws.title = "BOM"

        # 表头
        headers = ['序号', '零件名称', '数量', '配置', '材料', '质量(kg)']
        ws.append(headers)

        # 样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 数据
        for item in self.bom_items:
            ws.append([
                item.item_number,
                item.part_name,
                item.quantity,
                item.configuration,
                item.material if item.material else "-",
                f"{item.mass:.4f}" if item.mass > 0 else "-"
            ])

        # 列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 12

        # 统计
        ws_stats = wb.create_sheet("统计")
        ws_stats.append(['统计项', '数值'])
        ws_stats.append(['独特零件', len(self.bom_items)])
        ws_stats.append(['总数量', sum(item.quantity for item in self.bom_items)])

        for cell in ws_stats[1]:
            cell.fill = header_fill
            cell.font = header_font

        wb.save(output_path)
        print(f"✅ 已导出\n")

    def close_document(self):
        """关闭文档（改进版）"""
        if self.model_doc:
            try:
                print("关闭文档...")

                if self.sw_app:
                    try:
                        # 方法1: 通过路径关闭
                        doc_path = self.model_doc.GetPathName()
                        if doc_path:
                            self.sw_app.CloseDoc(doc_path)
                            print("✓ 文档已关闭")
                        else:
                            raise Exception("无法获取文档路径")

                    except Exception as e1:
                        # 方法2: 关闭所有文档
                        try:
                            print(f"   方法1失败: {e1}")
                            print("   尝试关闭所有文档...")
                            self.sw_app.CloseAllDocuments(True)  # True = 不保存更改
                            print("✓ 已关闭所有文档")
                        except Exception as e2:
                            print(f"   方法2失败: {e2}")

                self.model_doc = None

            except Exception as e:
                print(f"⚠ 关闭文档时出错: {e}")
                self.model_doc = None

    def quit_solidworks(self):
        """退出SolidWorks（改进版）"""
        if self.sw_app:
            try:
                # 只退出我们自己启动的实例
                if self.started_solidworks:
                    print("退出SolidWorks...")
                    try:
                        self.sw_app.ExitApp()
                        print("✓ SolidWorks已退出")
                    except Exception as e:
                        print(f"⚠ 退出失败: {e}")
                        # 强制结束进程（备用）
                        try:
                            import subprocess
                            subprocess.call(['taskkill', '/F', '/IM', 'SLDWORKS. exe'],
                                            stdout=subprocess.DEVNULL,
                                            stderr=subprocess.DEVNULL)
                            print("✓ 已强制结束SolidWorks进程")
                        except:
                            pass
                else:
                    print("保留现有SolidWorks实例")

            except Exception as e:
                print(f"⚠ 退出过程出错: {e}")

            finally:
                self.sw_app = None

        try:
            pythoncom.CoUninitialize()
        except:
            pass

    # ✅ 添加JSON和CSV导出方法
    def export_to_json(self, output_path: str):
        """导出为JSON"""
        import json
        from datetime import datetime

        print(f"📊 导出到JSON: {os.path.basename(output_path)}")

        bom_data = {
            'generated_at': datetime.now().isoformat(),
            'extraction_mode': 'auto',
            'total_items': len(self.bom_items),
            'total_quantity': sum(item.quantity for item in self.bom_items),
            'items': [item.to_dict() for item in self.bom_items]
        }

        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(bom_data, f, ensure_ascii=False, indent=2)

        print(f"✅ 已导出\n")

    def export_to_csv(self, output_path: str):
        """导出到CSV"""
        import csv

        print(f"📊 导出到CSV: {os.path.basename(output_path)}")

        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)

            # 表头
            writer.writerow(['序号', '零件名称', '数量', '配置', '材料', '质量(kg)'])

            # 数据
            for item in self.bom_items:
                writer.writerow([
                    item.item_number,
                    item.part_name,
                    item.quantity,
                    item.configuration,
                    item.material if item.material else "-",
                    f"{item.mass:.4f}" if item.mass > 0 else "-"
                ])

            # 统计
            writer.writerow([])
            writer.writerow(['统计信息', '', '', '', '', ''])
            writer.writerow(['独特零件', len(self.bom_items), '', '', '', ''])
            writer.writerow(['总数量', sum(item.quantity for item in self.bom_items), '', '', '', ''])

        print(f"✅ 已导出\n")

    def print_summary(self):
        """打印摘要"""
        print("=" * 70)
        print("📋 BOM摘要")
        print("=" * 70)

        total = sum(item.quantity for item in self.bom_items)

        print(f"独特零件: {len(self.bom_items)}")
        print(f"总数量: {total}")

        print("\n" + "-" * 70)
        print(f"{'序号':<6} {'零件名称': <30} {'数量':<6} {'材料':<15}")
        print("-" * 70)

        for item in self.bom_items[: 20]:
            mat = item.material[: 13] if item.material else "-"
            print(f"{item.item_number:<6} {item.part_name:<30} {item.quantity:<6} {mat:<15}")

        if len(self.bom_items) > 20:
            print(f"... 还有 {len(self.bom_items) - 20} 项")

        print("=" * 70 + "\n")

    def export_component_step(self, component_name: str, output_path: str) -> bool:
        """
        导出指定零件/子装配为STEP文件（基于SolidWorks 2022 API官方文档）

        参考文档:
        - IComponent2::GetReferencedModelDoc()
        - IComponent2::GetSuppression (属性)
        - IModelDoc2::SaveAs3()

        Args:
            component_name: 零件/子装配名称
            output_path:     输出STEP文件路径

        Returns:
            是否导出成功
        """
        if not self.sw_app or not self.model_doc:
            print(f"      ⚠ SolidWorks未初始化")
            return False

        try:
            # 获取配置管理器
            config_mgr = self.model_doc.ConfigurationManager
            if not config_mgr:
                print(f"      ⚠ 无法获取配置管理器")
                return False

            active_config = config_mgr.ActiveConfiguration
            if not active_config:
                print(f"      ⚠ 无法获取活动配置")
                return False

            root_comp = active_config.GetRootComponent3(True)
            if not root_comp:
                print(f"      ⚠ 无法获取根组件")
                return False

            # GetChildren 是属性，不是方法
            children = root_comp.GetChildren

            if not children or len(children) == 0:
                print(f"      ⚠ 装配中没有子组件")
                return False

            # 规范化组件名称用于匹配
            component_normalized = component_name.replace(' ', '_').lower()

            # 遍历子组件
            for child in children:
                try:
                    comp_name = child.Name2

                    # 获取基础名称（去掉实例编号）
                    base_name = comp_name.split('-')[0] if '-' in comp_name else comp_name
                    base_normalized = base_name.replace(' ', '_').lower()

                    # 检查是否匹配
                    if (component_normalized in base_normalized or
                            base_normalized.startswith(component_normalized)):

                        print(f"      ✓ 找到匹配组件: {comp_name}")

                        # ✅ 使用 GetReferencedModelDoc() 而不是 GetModelDoc2()
                        # 这个方法会自动处理轻量化组件
                        comp_model = child.GetReferencedModelDoc()

                        if not comp_model:
                            print(f"      ⚠ 无法获取组件引用文档（可能被抑制）")
                            continue

                        # 确保输出目录存在
                        output_dir = os.path.dirname(output_path)
                        if output_dir and not os.path.exists(output_dir):
                            os.makedirs(output_dir)

                        # ✅ 清除选择（如果方法存在）
                        try:
                            comp_model.ClearSelection2(True)
                        except:
                            pass

                        # ✅ 使用 SaveAs3 导出STEP
                        try:
                            success = comp_model.SaveAs3(
                                output_path,
                                0,  # swSaveAsCurrentVersion
                                0  # swSaveAsOptions_Silent
                            )

                            if success:
                                # 验证文件是否真的创建了
                                if os.path.exists(output_path):
                                    file_size = os.path.getsize(output_path)
                                    print(f"      ✅ STEP已导出: {os.path.basename(output_path)} ({file_size} bytes)")
                                    return True
                                else:
                                    print(f"      ⚠ SaveAs3成功但文件不存在")
                            else:
                                print(f"      ⚠ SaveAs3返回False")

                        except Exception as save_ex:
                            print(f"      ⚠ SaveAs3异常: {save_ex}")
                            continue

                except Exception as e:
                    print(f"      ⚠ 处理组件异常: {e}")
                    continue

            print(f"      ⚠ 未找到可导出的匹配组件:  {component_name}")
            return False

        except Exception as e:
            print(f"      ❌ 导出组件失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def export_all_components_to_step(self, output_dir: str) -> Dict[str, str]:
        """批量导出装配中所有零件为STEP文件"""
        if not self.sw_app or not self.model_doc:
            print("⚠ SolidWorks未初始化")
            return {}

        print(f"\n{'=' * 70}")
        print(f"📤 批量导出零件为STEP")
        print(f"{'=' * 70}\n")

        os.makedirs(output_dir, exist_ok=True)

        exported = {}
        exported_parts = set()

        try:
            config_mgr = self.model_doc.ConfigurationManager
            if not config_mgr:
                print("⚠ 无法获取配置管理器")
                return {}

            active_config = config_mgr.ActiveConfiguration
            if not active_config:
                print("⚠ 无法获取活动配置")
                return {}

            root_comp = active_config.GetRootComponent3(True)
            if not root_comp:
                print("⚠ 无法获取根组件")
                return {}

            children = root_comp.GetChildren

            if not children or len(children) == 0:
                print("⚠ 装配中没有子组件")
                return {}

            total = len(children)
            print(f"找到 {total} 个组件实例\n")

            for idx, child in enumerate(children, 1):
                try:
                    comp_name = child.Name2
                    base_name = comp_name.split('-')[0] if '-' in comp_name else comp_name
                    base_name_normalized = base_name.replace(' ', '_')

                    if base_name_normalized in exported_parts:
                        print(f"  [{idx}/{total}] ⊙ 跳过（已导出）:  {base_name}")
                        continue

                    print(f"  [{idx}/{total}] 📦 处理: {base_name}")

                    # 获取组件路径
                    comp_path = child.GetPathName  # 属性

                    if not comp_path:
                        print(f"         ⚠ 无法获取组件路径")
                        continue

                    comp_doc = None

                    # ✅ 关键修复：在 Python win32com 中，ref 参数（Errors, Warnings）应该省略
                    # OpenDoc6 只传前4个参数，Errors 和 Warnings 会作为返回值的一部分

                    # 尝试作为零件打开
                    try:
                        comp_doc = self.sw_app.OpenDoc6(
                            comp_path,  # FileName
                            1,  # Type:  swDocPART
                            1,  # Options:  swOpenDocOptions_Silent
                            ""  # Configuration
                            # ← Errors 和 Warnings 参数省略！
                        )
                    except Exception as e:
                        print(f"         ⚠ 作为零件打开失败:  {e}")

                    # 如果失败，尝试作为装配打开
                    if not comp_doc:
                        try:
                            comp_doc = self.sw_app.OpenDoc6(
                                comp_path,
                                2,  # Type: swDocASSEMBLY
                                1,
                                ""
                            )
                        except Exception as e:
                            print(f"         ⚠ 作为装配打开失败: {e}")

                    if not comp_doc:
                        print(f"         ⚠ 无法打开组件文档")
                        continue

                    # 生成STEP文件名
                    step_filename = f"{base_name_normalized}_bom_v1.step"
                    step_path = os.path.join(output_dir, step_filename)

                    # 清除选择
                    try:
                        comp_doc.ClearSelection2(True)
                    except:
                        pass

                    # 导出STEP
                    success = comp_doc.SaveAs3(
                        step_path,
                        0,
                        0
                    )

                    # 关闭文档
                    try:
                        self.sw_app.CloseDoc(comp_path)
                    except:
                        pass

                    if success and os.path.exists(step_path):
                        file_size = os.path.getsize(step_path)
                        print(f"         ✅ 已导出: {step_filename} ({file_size: ,} bytes)")
                        exported[base_name_normalized] = step_path
                        exported_parts.add(base_name_normalized)
                    else:
                        print(f"         ⚠ 导出失败")

                except Exception as e:
                    print(f"         ⚠ 异常:  {e}")
                    continue

            print(f"\n{'=' * 70}")
            print(f"✅ 批量导出完成: {len(exported)} 个零件")
            print(f"{'=' * 70}\n")

            return exported

        except Exception as e:
            print(f"❌ 批量导出失败: {e}")
            import traceback
            traceback.print_exc()
            return exported

    def run_macro(self, macro_path: str, module_name: str, procedure_name: str) -> bool:
        """
        运行SolidWorks宏（基于SolidWorks 2022 API官方文档）

        官方文档：
        bool RunMacro2(
            string FilePathName,
            string ModuleName,
            string ProcedureName,
            int Options,
            ref int Error
        )

        Args:
            macro_path: 宏文件路径 (. swp)
            module_name: 模块名称（通常是 "Macro1"）
            procedure_name: 过程名称（如 "main"）

        Returns:
            是否成功执行
        """
        if not self.sw_app:
            print("⚠ SolidWorks未初始化")
            return False

        if not os.path.exists(macro_path):
            print(f"⚠ 宏文件不存在: {macro_path}")
            return False

        try:
            print(f"   🔧 运行宏: {os.path.basename(macro_path)}")
            print(f"      模块: {module_name}")
            print(f"      过程: {procedure_name}")

            # ✅ 根据官方文档，RunMacro2 需要5个参数
            # 在 Python win32com 中，ref 参数需要使用 VARIANT 包装

            import pythoncom
            from win32com.client import VARIANT

            # 创建一个 VARIANT 用于接收 Error 参数
            error_code = VARIANT(pythoncom.VT_I4 | pythoncom.VT_BYREF, 0)

            # Options 枚举值：
            # swRunMacroDefault = 0
            # swRunMacroUnloadAfterRun = 1
            options = 1  # swRunMacroUnloadAfterRun（执行后卸载，避免内存问题）

            success = self.sw_app.RunMacro2(
                macro_path,  # FilePathName
                module_name,  # ModuleName
                procedure_name,  # ProcedureName
                options,  # Options
                error_code  # Error (ref parameter)
            )

            if success:
                print(f"   ✓ 宏执行成功")
                return True
            else:
                error_val = error_code.value if hasattr(error_code, 'value') else 0
                print(f"   ⚠ 宏执行失败，错误码: {error_val}")
                return False

        except Exception as e:
            print(f"   ❌ 运行宏异常: {e}")
            import traceback
            traceback.print_exc()
            return False